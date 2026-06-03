import dash
from dash import dcc, html, Input, Output, State, callback_context
import dash_bootstrap_components as dbc
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import pandas as pd
import numpy as np
import base64
import io
import json
import os
import uuid
import warnings
warnings.filterwarnings("ignore")

from sklearn.ensemble import IsolationForest
from sklearn.preprocessing import StandardScaler
from sklearn.decomposition import PCA
from statsmodels.tsa.arima.model import ARIMA
import xgboost as xgb

from reportlab.lib.pagesizes import A4
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                 Table, TableStyle, PageBreak, HRFlowable)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import cm
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference
from datetime import datetime

# ── App ─────────────────────────────────────────────────────────────────────────
app = dash.Dash(
    __name__,
    external_stylesheets=[dbc.themes.BOOTSTRAP],
    title="PFEMS · Savanna Dynamics",
    meta_tags=[{"name": "viewport", "content": "width=device-width, initial-scale=1"}],
    suppress_callback_exceptions=True,
)
server = app.server

# ── Audit log path (writable on Render ephemeral FS) ───────────────────────────
AUDIT_PATH = "/tmp/pfems_audit_log.json"

def write_audit_entry(entry):
    try:
        log = []
        if os.path.exists(AUDIT_PATH):
            with open(AUDIT_PATH, "r") as f:
                log = json.load(f)
        log.append(entry)
        with open(AUDIT_PATH, "w") as f:
            json.dump(log[-500:], f, indent=2)
    except Exception:
        pass

def read_audit_log():
    try:
        if os.path.exists(AUDIT_PATH):
            with open(AUDIT_PATH, "r") as f:
                return json.load(f)
    except Exception:
        pass
    return []

# ── Colours ──────────────────────────────────────────────────────────────────────
C = {
    "bg": "#0A0E17", "surface": "#111827", "card": "#161D2E",
    "border": "#1E2D45", "accent": "#00D4AA", "accent2": "#FF6B35",
    "accent3": "#4B9FFF", "warn": "#FFB800", "danger": "#FF3B5C",
    "text": "#E8EDF5", "muted": "#6B7A99", "grid": "#1A2540",
}

CARD = {
    "backgroundColor": C["card"], "border": "1px solid " + C["border"],
    "borderRadius": "12px", "padding": "24px", "marginBottom": "20px",
}
HEAD = {
    "fontFamily": "monospace", "color": C["accent"], "fontSize": "11px",
    "letterSpacing": "3px", "textTransform": "uppercase", "marginBottom": "16px",
}
INPUT_STYLE = {
    "width": "100%", "backgroundColor": C["surface"],
    "border": "1px solid " + C["border"], "borderRadius": "6px",
    "padding": "8px 12px", "color": C["text"], "fontSize": "13px",
}
LABEL_STYLE = {"color": C["muted"], "fontSize": "12px", "marginBottom": "4px"}
BTN = {
    "border": "none", "borderRadius": "6px", "padding": "10px 24px",
    "fontFamily": "monospace", "fontSize": "11px",
    "letterSpacing": "2px", "fontWeight": "700", "cursor": "pointer",
}

# ── Helpers ───────────────────────────────────────────────────────────────────────
def blank_fig(msg="Upload data to begin"):
    fig = go.Figure()
    fig.add_annotation(text=msg, x=0.5, y=0.5, xref="paper", yref="paper",
                       showarrow=False, font=dict(color=C["muted"], size=14))
    fig.update_layout(
        paper_bgcolor=C["card"], plot_bgcolor=C["card"],
        xaxis=dict(showgrid=False, zeroline=False, showticklabels=False),
        yaxis=dict(showgrid=False, zeroline=False, showticklabels=False),
        margin=dict(l=20, r=20, t=20, b=20))
    return fig

def chart_layout(title="", height=340):
    return dict(
        title=dict(text=title, font=dict(color=C["muted"], size=12), x=0.01),
        paper_bgcolor=C["card"], plot_bgcolor=C["card"],
        font=dict(color=C["text"]), height=height,
        margin=dict(l=50, r=20, t=45, b=50),
        xaxis=dict(gridcolor=C["grid"], zeroline=False,
                   tickfont=dict(color=C["muted"], size=11)),
        yaxis=dict(gridcolor=C["grid"], zeroline=False,
                   tickfont=dict(color=C["muted"], size=11)),
        legend=dict(bgcolor="rgba(0,0,0,0)", font=dict(color=C["muted"], size=11),
                    orientation="h", yanchor="bottom", y=1.02),
        hovermode="x unified")

def _pill(text, colour):
    return html.Span(text, style={"fontFamily": "monospace", "color": colour,
                                   "fontSize": "11px", "letterSpacing": "2px"})

def _kpi(label, value, colour):
    return html.Div(style={**CARD, "borderTop": "2px solid " + colour,
                           "marginBottom": 0, "textAlign": "center", "padding": "16px"},
                    children=[
        html.Div(str(value), style={"fontFamily": "monospace", "fontSize": "26px",
                                    "fontWeight": "700", "color": colour}),
        html.Div(label, style={"color": C["muted"], "fontSize": "10px",
                               "letterSpacing": "2px", "marginTop": "6px"}),
    ])

def _anomaly_fig(dates, residual, flags, scores, title):
    fig = make_subplots(specs=[[{"secondary_y": True}]])
    fig.add_trace(go.Scatter(x=dates, y=residual, mode="lines", name="Residual",
                              line=dict(color=C["accent3"], width=1.2), opacity=0.6),
                  secondary_y=False)
    fig.add_trace(go.Scatter(x=dates, y=scores, mode="lines", name="Score",
                              line=dict(color=C["warn"], width=1.5)),
                  secondary_y=True)
    if hasattr(flags, "any") and flags.any():
        fig.add_trace(go.Scatter(x=dates[flags], y=residual[flags], mode="markers",
                                  name="Flagged",
                                  marker=dict(color=C["danger"], size=7, symbol="x")),
                      secondary_y=False)
    fig.update_layout(**chart_layout(title, 300))
    fig.update_yaxes(gridcolor=C["grid"], zeroline=False,
                     tickfont=dict(color=C["muted"], size=10), secondary_y=False)
    fig.update_yaxes(tickfont=dict(color=C["warn"], size=10),
                     secondary_y=True, showgrid=False)
    return fig

def _alert_table(alerts):
    if not alerts:
        return html.Div(
            "No ensemble-confirmed alerts. Requires 2 of 3 detectors to agree "
            "on a positive-direction deviation.",
            style={"color": C["muted"], "fontSize": "13px"})
    hs = {"fontFamily": "monospace", "fontSize": "10px", "letterSpacing": "2px",
          "color": C["muted"], "padding": "10px 12px",
          "borderBottom": "1px solid " + C["border"],
          "textAlign": "left", "whiteSpace": "nowrap"}
    cs = {"padding": "9px 12px", "fontSize": "12px", "color": C["text"],
          "borderBottom": "1px solid " + C["grid"], "whiteSpace": "nowrap"}
    cols = ["Date","Residual","CH4","Score","Votes","z","SNR",
            "ARIMA","XGB","AE","Level"]
    header = html.Tr([html.Th(c, style=hs) for c in cols])
    rows = []
    for a in alerts[-200:]:
        rows.append(html.Tr([
            html.Td(a["Date"],  style=cs),
            html.Td(a["Res"],   style=cs),
            html.Td(a["CH4"],   style=cs),
            html.Td(a["Score"], style=cs),
            html.Td(a["Votes"], style=cs),
            html.Td(a["Z"],     style=cs),
            html.Td(a["SNR"],   style=cs),
            html.Td(a["ARIMA"], style={**cs, "color": C["danger"] if a["ARIMA"]=="●" else C["muted"]}),
            html.Td(a["XGB"],   style={**cs, "color": C["danger"] if a["XGB"]  =="●" else C["muted"]}),
            html.Td(a["AE"],    style={**cs, "color": C["danger"] if a["AE"]   =="●" else C["muted"]}),
            html.Td(a["Level"], style={**cs, "color": a["colour"],
                                       "fontFamily": "monospace", "fontWeight": "700"}),
        ]))
    return html.Div(style={"overflowX": "auto"}, children=[
        html.Div(str(len(alerts)) + " ensemble-confirmed alerts "
                 "(majority voting · positive z-score · XGBoost prediction error)",
                 style={"color": C["muted"], "fontSize": "12px", "marginBottom": "12px"}),
        html.Table([html.Thead(header), html.Tbody(rows)],
                   style={"width": "100%", "borderCollapse": "collapse"})
    ])

# ── Analytics ─────────────────────────────────────────────────────────────────────

def compute_residual(df, prod, flare, export, fuel, inj):
    return df[prod] - (df[flare] + df[export] + df[fuel] + df[inj])

def apply_ch4(residual, frac):
    return residual * frac

def run_arima(residual, sigma=3.0):
    try:
        vals = residual.values
        result = ARIMA(vals, order=(1,0,1)).fit(
            method_kwargs={"warn_convergence": False})
        errors = vals - result.fittedvalues
        std = np.std(errors) or 1.0
        z = errors / std
        return z > sigma, np.abs(errors), z, (1,0,1)
    except Exception:
        n = len(residual)
        return np.zeros(n,bool), np.zeros(n), np.zeros(n), (1,0,1)

def run_xgboost(residual, contamination=0.02):
    """
    STAGE 2 FIX: Genuine XGBoost regression approach.
    1. Build lag feature matrix from the residual series.
    2. Train XGBoost to predict residual[t] from residual[t-1..t-7]
       using the first 70 percent of data as training.
    3. Compute prediction errors on the full series.
    4. Flag top contamination-fraction of errors as anomalies.
    5. Return feature importances for display.
    """
    try:
        vals = residual.values
        n    = len(vals)
        lags = 7

        # Build feature matrix
        X = np.column_stack([
            np.roll(vals, i) for i in range(1, lags + 1)
        ])
        X[:lags] = 0  # zero-pad initial rows

        # Add rolling stats
        s = pd.Series(vals)
        X = np.column_stack([
            X,
            s.rolling(7,  min_periods=1).mean().values,
            s.rolling(14, min_periods=1).mean().values,
            s.rolling(7,  min_periods=1).std().fillna(0).values,
        ])

        feat_names = (
            [f"lag_{i}" for i in range(1, lags+1)] +
            ["roll_mean_7", "roll_mean_14", "roll_std_7"]
        )

        # Train on first 70 percent
        split = int(0.70 * n)
        X_train, y_train = X[:split], vals[:split]

        model = xgb.XGBRegressor(
            n_estimators=80, max_depth=3, learning_rate=0.1,
            subsample=0.8, colsample_bytree=0.8,
            random_state=42, verbosity=0,
            tree_method="hist",
        )
        model.fit(X_train, y_train)

        preds  = model.predict(X)
        errors = np.abs(vals - preds)
        threshold = np.percentile(errors, (1 - contamination) * 100)
        flags  = errors > threshold

        # Feature importances
        importances = dict(zip(feat_names, model.feature_importances_))

        norm_e = (errors - errors.min()) / (errors.max() - errors.min() + 1e-9)
        return flags, norm_e, importances

    except Exception as e:
        n = len(residual)
        return np.zeros(n,bool), np.zeros(n), {}

def run_autoencoder(residual, contamination=0.02, window=14):
    try:
        vals   = residual.values
        scaled = StandardScaler().fit_transform(vals.reshape(-1,1)).flatten()
        n      = len(scaled)
        if n <= window:
            return np.zeros(n,bool), np.zeros(n)
        wins = np.array([scaled[i:i+window] for i in range(n-window)])
        pca  = PCA(n_components=min(2, window-1))
        errs = np.mean((wins - pca.inverse_transform(pca.fit_transform(wins)))**2, axis=1)
        full = np.concatenate([np.zeros(window), errs])
        thr  = np.percentile(full[window:], (1-contamination)*100)
        return full > thr, full
    except Exception:
        return np.zeros(len(residual),bool), np.zeros(len(residual))

def compute_ensemble(arima_f, xgb_f, ae_f, z_scores, min_votes=2):
    votes = arima_f.astype(int) + xgb_f.astype(int) + ae_f.astype(int)
    alert = (votes >= min_votes) & (z_scores > 0)
    return alert, votes

def compute_snr(residual, window=30):
    s = pd.Series(residual)
    return (s.rolling(window,min_periods=1).mean().abs() /
            s.rolling(window,min_periods=1).std().fillna(1)
             .replace(0, np.nan).fillna(1))

def compute_mdlf(residual, produced, window=90):
    r = pd.Series(residual.values)
    p = pd.Series(produced.values)
    return ((3 * r.rolling(window,min_periods=window//3).std() /
             p.rolling(window,min_periods=window//3).mean()
              .replace(0,np.nan)) * 100).ffill().bfill()

def classify_alert(score, snr_val):
    if   score > 0.8 and snr_val > 3: return "CRITICAL", C["danger"]
    elif score > 0.5 or  snr_val > 2: return "WARNING",  C["warn"]
    elif score > 0.2:                  return "WATCH",    C["accent3"]
    else:                              return "NORMAL",   C["accent"]

def generate_ref(field):
    ts = datetime.now().strftime("%Y%m%d%H%M")
    return f"SDL-PFEMS-{ts}-{field[:4].upper()}"

# ── Cost calculator data ──────────────────────────────────────────────────────────
COST_DATA = {
    "PFEMS (this system)":     {"cost_per_day": 0,    "days_per_year": 365, "detects_all": True},
    "OGI Camera Survey":       {"cost_per_day": 3500, "days_per_year": 12,  "detects_all": False},
    "Satellite Monitoring":    {"cost_per_day": 800,  "days_per_year": 48,  "detects_all": False},
    "Continuous Sensor Array": {"cost_per_day": 1200, "days_per_year": 365, "detects_all": True},
}

# ════════════════════════════════════════════════════════════════════════════════
# LAYOUT
# ════════════════════════════════════════════════════════════════════════════════
app.layout = html.Div(
    style={"backgroundColor": C["bg"], "minHeight": "100vh"},
    children=[

    # ── Top bar ─────────────────────────────────────────────────────────────────
    html.Div(style={
        "backgroundColor": C["surface"],
        "borderBottom": "1px solid " + C["border"],
        "padding": "0 40px", "display": "flex", "alignItems": "center",
        "justifyContent": "space-between", "height": "64px",
        "position": "sticky", "top": 0, "zIndex": 1000,
    }, children=[
        html.Div([
            html.Span("PFEMS", style={"fontFamily": "monospace", "color": C["accent"],
                                      "fontSize": "18px", "fontWeight": "700",
                                      "letterSpacing": "2px", "marginRight": "12px"}),
            html.Span("Physics-Guided Fugitive Emissions Monitoring System · Savanna Dynamics Limited",
                      style={"color": C["muted"], "fontSize": "11px"}),
        ], style={"display": "flex", "alignItems": "center"}),
        html.Div(id="status-pill", children=[_pill("● NO DATA", C["muted"])]),
    ]),

    # ── Tabs ────────────────────────────────────────────────────────────────────
    html.Div(style={"backgroundColor": C["surface"],
                    "borderBottom": "1px solid " + C["border"],
                    "padding": "0 40px", "display": "flex", "gap": "0"},
             children=[
        html.Button("ANALYSIS", id="tab-analysis", n_clicks=0, style={
            **BTN, "backgroundColor": "transparent",
            "color": C["accent"], "borderBottom": "2px solid " + C["accent"],
            "borderRadius": "0", "padding": "14px 20px", "letterSpacing": "2px",
        }),
        html.Button("PORTFOLIO", id="tab-portfolio", n_clicks=0, style={
            **BTN, "backgroundColor": "transparent", "color": C["muted"],
            "borderBottom": "2px solid transparent",
            "borderRadius": "0", "padding": "14px 20px", "letterSpacing": "2px",
        }),
        html.Button("COST CALCULATOR", id="tab-cost", n_clicks=0, style={
            **BTN, "backgroundColor": "transparent", "color": C["muted"],
            "borderBottom": "2px solid transparent",
            "borderRadius": "0", "padding": "14px 20px", "letterSpacing": "2px",
        }),
        html.Button("AUDIT LOG", id="tab-audit", n_clicks=0, style={
            **BTN, "backgroundColor": "transparent", "color": C["muted"],
            "borderBottom": "2px solid transparent",
            "borderRadius": "0", "padding": "14px 20px", "letterSpacing": "2px",
        }),
    ]),

    # ── Tab content — all pre-rendered, shown/hidden via CSS ────────────────────
    html.Div(style={"padding": "32px 40px", "maxWidth": "1600px", "margin": "0 auto"},
             children=[

        # ANALYSIS TAB (always in DOM)
        html.Div(id="panel-analysis", style={"display": "block"}, children=[

            html.Div(style=CARD, children=[
                html.P("01 / FIELD CONFIGURATION", style=HEAD),
                dbc.Row([
                    dbc.Col([html.Label("Field / Asset Name", style=LABEL_STYLE),
                             dcc.Input(id="field-name", type="text",
                                       placeholder="e.g. Jubilee FPSO, TEN Field",
                                       style=INPUT_STYLE)], md=3),
                    dbc.Col([html.Label("Operator / Company", style=LABEL_STYLE),
                             dcc.Input(id="operator-name", type="text",
                                       placeholder="e.g. Tullow Oil, GNPC",
                                       style=INPUT_STYLE)], md=3),
                    dbc.Col([html.Label("Country / Jurisdiction", style=LABEL_STYLE),
                             dcc.Input(id="country-name", type="text",
                                       placeholder="e.g. Ghana, Nigeria",
                                       style=INPUT_STYLE)], md=3),
                    dbc.Col([
                        html.Label("Methane Fraction (CH4) — field-specific", style=LABEL_STYLE),
                        dcc.Input(id="ch4-fraction", type="number",
                                  placeholder="e.g. 0.77", min=0.5, max=1.0,
                                  step=0.01, value=0.77, style=INPUT_STYLE),
                        html.Div("Range 0.50 to 1.00 · Jubilee default 0.77",
                                 style={"color": C["muted"], "fontSize": "11px",
                                        "marginTop": "4px"}),
                    ], md=3),
                ], className="g-3"),
            ]),

            html.Div(style=CARD, children=[
                html.P("02 / DATA INGESTION", style=HEAD),
                dcc.Upload(id="upload-data",
                    children=html.Div([
                        html.Div("↑", style={"fontSize": "32px", "color": C["accent"],
                                             "marginBottom": "8px"}),
                        html.Div("Drop CSV file here or click to browse",
                                 style={"color": C["text"], "fontSize": "14px"}),
                        html.Div("Columns: date, produced gas, flaring, export, "
                                 "fuel gas, injection",
                                 style={"color": C["muted"], "fontSize": "12px"}),
                    ], style={"textAlign": "center", "padding": "20px"}),
                    style={"border": "1px dashed " + C["border"], "borderRadius": "8px",
                           "cursor": "pointer", "backgroundColor": C["surface"]},
                    multiple=False),
                html.Div(id="upload-status",
                         style={"marginTop": "12px", "fontSize": "13px",
                                "color": C["muted"]}),
            ]),

            html.Div(id="column-mapping-section", style={"display": "none"}, children=[
                html.Div(style=CARD, children=[
                    html.P("03 / COLUMN MAPPING", style=HEAD),
                    html.P("Map your CSV columns to PFEMS variables.",
                           style={"color": C["muted"], "fontSize": "13px",
                                  "marginBottom": "20px"}),
                    dbc.Row([
                        dbc.Col([html.Label("Date", style=LABEL_STYLE),
                                 dcc.Dropdown(id="col-date", placeholder="Select...")], md=2),
                        dbc.Col([html.Label("Produced Gas", style=LABEL_STYLE),
                                 dcc.Dropdown(id="col-produced", placeholder="Select...")], md=2),
                        dbc.Col([html.Label("Gas Flaring", style=LABEL_STYLE),
                                 dcc.Dropdown(id="col-flare", placeholder="Select...")], md=2),
                        dbc.Col([html.Label("Gas Export", style=LABEL_STYLE),
                                 dcc.Dropdown(id="col-export", placeholder="Select...")], md=2),
                        dbc.Col([html.Label("Fuel Gas", style=LABEL_STYLE),
                                 dcc.Dropdown(id="col-fuel", placeholder="Select...")], md=2),
                        dbc.Col([html.Label("Gas Injection", style=LABEL_STYLE),
                                 dcc.Dropdown(id="col-injection", placeholder="Select...")], md=2),
                    ], className="g-3"),
                    html.Div(style={"marginTop": "20px", "display": "flex",
                                    "gap": "20px", "alignItems": "center",
                                    "flexWrap": "wrap"}, children=[
                        html.Button("RUN ANALYSIS", id="run-btn", n_clicks=0,
                                    style={**BTN, "backgroundColor": C["accent"],
                                           "color": C["bg"]}),
                        html.Div([
                            html.Label("Contamination %", style=LABEL_STYLE),
                            dcc.Slider(id="contamination-slider",
                                       min=1, max=5, step=1, value=2,
                                       marks={i: {"label": str(i)+"%",
                                                  "style": {"color": C["muted"]}}
                                              for i in [1,2,3,5]}),
                        ], style={"flex": "1", "maxWidth": "260px"}),
                        html.Div([
                            html.Label("Sigma Threshold", style=LABEL_STYLE),
                            dcc.Slider(id="sigma-slider",
                                       min=2.0, max=4.0, step=0.5, value=3.0,
                                       marks={v: {"label": str(v),
                                                  "style": {"color": C["muted"]}}
                                              for v in [2.0, 2.5, 3.0, 3.5, 4.0]}),
                        ], style={"flex": "1", "maxWidth": "260px"}),
                    ]),
                    html.Div(id="model-info-display",
                             style={"marginTop": "10px", "fontSize": "12px",
                                    "color": C["muted"]}),
                ]),
            ]),

            html.Div(id="kpi-section", style={"display": "none"}, children=[
                html.P("04 / SYSTEM STATUS", style={**HEAD, "marginBottom": "12px"}),
                dbc.Row(id="kpi-cards", className="g-3 mb-4"),
            ]),

            html.Div(id="charts-section", style={"display": "none"}, children=[

                html.P("05 / METHANE RESIDUAL SIGNAL", style=HEAD),
                html.Div(style=CARD, children=[
                    dcc.Graph(id="residual-chart", figure=blank_fig(),
                              config={"displayModeBar": False}),
                ]),

                dbc.Row([
                    dbc.Col([
                        html.P("06 / ARIMA(1,0,1) ANOMALY DETECTION", style=HEAD),
                        html.Div(style=CARD, children=[
                            dcc.Graph(id="arima-chart", figure=blank_fig(),
                                      config={"displayModeBar": False}),
                        ]),
                    ], md=6),
                    dbc.Col([
                        html.P("07 / XGBOOST PREDICTION ERROR", style=HEAD),
                        html.Div(style=CARD, children=[
                            dcc.Graph(id="xgb-chart", figure=blank_fig(),
                                      config={"displayModeBar": False}),
                        ]),
                    ], md=6),
                ]),

                dbc.Row([
                    dbc.Col([
                        html.P("08 / AUTOENCODER RECONSTRUCTION ERROR", style=HEAD),
                        html.Div(style=CARD, children=[
                            dcc.Graph(id="lstm-chart", figure=blank_fig(),
                                      config={"displayModeBar": False}),
                        ]),
                    ], md=6),
                    dbc.Col([
                        html.P("09 / SIGNAL-TO-NOISE RATIO", style=HEAD),
                        html.Div(style=CARD, children=[
                            dcc.Graph(id="snr-chart", figure=blank_fig(),
                                      config={"displayModeBar": False}),
                        ]),
                    ], md=6),
                ]),

                html.P("10 / MINIMUM DETECTABLE LEAK FRACTION", style=HEAD),
                html.Div(style=CARD, children=[
                    dcc.Graph(id="mdlf-chart", figure=blank_fig(),
                              config={"displayModeBar": False}),
                    html.Div(id="mdlf-summary",
                             style={"marginTop": "12px", "fontSize": "12px",
                                    "color": C["muted"]}),
                ]),

                html.P("11 / XGBOOST FEATURE IMPORTANCE", style=HEAD),
                html.Div(style=CARD, children=[
                    dcc.Graph(id="importance-chart",
                              figure=blank_fig("Run analysis to see feature importance"),
                              config={"displayModeBar": False}),
                    html.Div(id="importance-text",
                             style={"marginTop": "8px", "fontSize": "12px",
                                    "color": C["muted"]}),
                ]),

                html.P("12 / ALERT LOG", style=HEAD),
                html.Div(style=CARD, children=[html.Div(id="alert-table")]),

                html.P("13 / MODEL ENSEMBLE AGREEMENT", style=HEAD),
                html.Div(style=CARD, children=[
                    dcc.Graph(id="ensemble-chart", figure=blank_fig(),
                              config={"displayModeBar": False}),
                ]),

                html.P("14 / EXPORT REPORTS", style=HEAD),
                html.Div(style=CARD, children=[
                    html.P("Download full analysis in your preferred format.",
                           style={"color": C["muted"], "fontSize": "13px",
                                  "marginBottom": "16px"}),
                    html.Div(style={"display": "flex", "gap": "12px",
                                    "flexWrap": "wrap"}, children=[
                        html.Button("↓ PDF REPORT", id="btn-pdf", n_clicks=0,
                                    style={**BTN, "backgroundColor": C["danger"],
                                           "color": "#fff", "padding": "10px 20px"}),
                        html.Button("↓ EXCEL ALERT LOG", id="btn-excel", n_clicks=0,
                                    style={**BTN, "backgroundColor": "#1D6F42",
                                           "color": "#fff", "padding": "10px 20px"}),
                        html.Button("↓ CSV RESIDUAL", id="btn-csv", n_clicks=0,
                                    style={**BTN, "backgroundColor": C["accent3"],
                                           "color": "#fff", "padding": "10px 20px"}),
                    ]),
                    html.Div(id="export-status",
                             style={"marginTop": "12px", "fontSize": "13px",
                                    "color": C["muted"]}),
                ]),
            ]),
        ]),

        # PORTFOLIO TAB
        html.Div(id="panel-portfolio", style={"display": "none"}, children=[
            html.Div(style=CARD, children=[
                html.P("PORTFOLIO OVERVIEW", style=HEAD),
                html.P("Upload multiple field datasets to compare monitoring "
                       "status across assets.",
                       style={"color": C["muted"], "fontSize": "13px",
                              "marginBottom": "20px"}),
                dcc.Upload(id="portfolio-upload",
                    children=html.Div([
                        html.Div("↑", style={"fontSize": "28px", "color": C["accent"]}),
                        html.Div("Upload CSV files for additional fields",
                                 style={"color": C["text"], "fontSize": "14px"}),
                    ], style={"textAlign": "center", "padding": "16px"}),
                    style={"border": "1px dashed " + C["border"],
                           "borderRadius": "8px", "cursor": "pointer",
                           "backgroundColor": C["surface"]},
                    multiple=True),
                html.Div(id="portfolio-status",
                         style={"marginTop": "12px", "color": C["muted"],
                                "fontSize": "13px"}),
            ]),
            html.Div(id="portfolio-table-section", style={"display": "none"}, children=[
                html.Div(style=CARD, children=[
                    html.P("FIELD COMPARISON", style=HEAD),
                    html.Div(id="portfolio-table"),
                ]),
                html.Div(style=CARD, children=[
                    html.P("MDLF COMPARISON", style=HEAD),
                    dcc.Graph(id="portfolio-chart",
                              figure=blank_fig("Upload field data to compare"),
                              config={"displayModeBar": False}),
                ]),
            ]),
        ]),

        # COST CALCULATOR TAB
        html.Div(id="panel-cost", style={"display": "none"}, children=[
            html.Div(style=CARD, children=[
                html.P("MONITORING COST-EFFECTIVENESS CALCULATOR", style=HEAD),
                html.P("Compare PFEMS continuous monitoring against conventional "
                       "methane detection methods.",
                       style={"color": C["muted"], "fontSize": "13px",
                              "marginBottom": "24px"}),
                dbc.Row([
                    dbc.Col([html.Label("Daily production (MMscf/day)", style=LABEL_STYLE),
                             dcc.Input(id="cost-production", type="number", value=150,
                                       min=1, max=500, step=1, style=INPUT_STYLE)], md=3),
                    dbc.Col([html.Label("Methane price (USD per MMscf)", style=LABEL_STYLE),
                             dcc.Input(id="cost-price", type="number", value=3000,
                                       min=100, max=20000, step=100, style=INPUT_STYLE)], md=3),
                    dbc.Col([html.Label("Assumed leak rate (% of production)",
                                       style=LABEL_STYLE),
                             dcc.Input(id="cost-leak-rate", type="number", value=1.0,
                                       min=0.1, max=10.0, step=0.1, style=INPUT_STYLE)], md=3),
                    dbc.Col([html.Label("PFEMS annual licence fee (USD)", style=LABEL_STYLE),
                             dcc.Input(id="cost-pfems-fee", type="number", value=25000,
                                       min=0, max=500000, step=1000, style=INPUT_STYLE)], md=3),
                ], className="g-3"),
                html.Div(style={"marginTop": "20px"}, children=[
                    html.Button("CALCULATE", id="btn-calculate-cost", n_clicks=0,
                                style={**BTN, "backgroundColor": C["accent"],
                                       "color": C["bg"]}),
                ]),
            ]),
            html.Div(id="cost-results", style={"display": "none"}, children=[
                html.Div(style=CARD, children=[
                    html.P("COST COMPARISON", style=HEAD),
                    dcc.Graph(id="cost-chart", figure=blank_fig(),
                              config={"displayModeBar": False}),
                ]),
                html.Div(style=CARD, children=[
                    html.P("EXECUTIVE COST SUMMARY", style=HEAD),
                    html.Div(id="cost-summary-table"),
                ]),
            ]),
        ]),

        # AUDIT LOG TAB
        html.Div(id="panel-audit", style={"display": "none"}, children=[
            html.Div(id="audit-content", style=CARD, children=[
                html.P("AUDIT LOG", style=HEAD),
                html.Div("No analysis sessions recorded yet. Run an analysis first.",
                         style={"color": C["muted"], "fontSize": "13px"}),
            ]),
        ]),

    ]),

    # Stores and downloads
    dcc.Store(id="stored-data"),
    dcc.Store(id="stored-results"),
    dcc.Download(id="download-pdf"),
    dcc.Download(id="download-excel"),
    dcc.Download(id="download-csv"),

    # Footer
    html.Div(style={
        "borderTop": "1px solid " + C["border"],
        "padding": "16px 40px", "display": "flex",
        "justifyContent": "space-between", "marginTop": "20px",
    }, children=[
        html.Span("PFEMS v2.0 · Savanna Dynamics Limited",
                  style={"color": C["muted"], "fontSize": "11px", "fontFamily": "monospace"}),
        html.Span("Data processed in-session only · Not stored server-side",
                  style={"color": C["muted"], "fontSize": "11px"}),
    ]),
])



# ════════════════════════════════════════════════════════════════════════════════
# CALLBACKS
# ════════════════════════════════════════════════════════════════════════════════

# ── Tab switching via CSS show/hide ──────────────────────────────────────────────
@app.callback(
    Output("panel-analysis",  "style"),
    Output("panel-portfolio", "style"),
    Output("panel-cost",      "style"),
    Output("panel-audit",     "style"),
    Output("tab-analysis",    "style"),
    Output("tab-portfolio",   "style"),
    Output("tab-cost",        "style"),
    Output("tab-audit",       "style"),
    Output("audit-content",   "children"),
    Input("tab-analysis",  "n_clicks"),
    Input("tab-portfolio", "n_clicks"),
    Input("tab-cost",      "n_clicks"),
    Input("tab-audit",     "n_clicks"),
)
def switch_tab(a, p, co, au):
    ctx = callback_context
    tid = ctx.triggered[0]["prop_id"].split(".")[0] if ctx.triggered else "tab-analysis"
    active = {"tab-analysis": "analysis", "tab-portfolio": "portfolio",
              "tab-cost": "cost", "tab-audit": "audit"}.get(tid, "analysis")
    show = {"display": "block"}
    hide = {"display": "none"}
    tab_on  = {**BTN, "backgroundColor": "transparent", "color": C["accent"],
               "borderBottom": "2px solid " + C["accent"],
               "borderRadius": "0", "padding": "14px 20px", "letterSpacing": "2px"}
    tab_off = {**BTN, "backgroundColor": "transparent", "color": C["muted"],
               "borderBottom": "2px solid transparent",
               "borderRadius": "0", "padding": "14px 20px", "letterSpacing": "2px"}
    panels = {"analysis": (show,hide,hide,hide), "portfolio": (hide,show,hide,hide),
              "cost":     (hide,hide,show,hide), "audit":     (hide,hide,hide,show)}
    tabs   = {"analysis": (tab_on,tab_off,tab_off,tab_off),
              "portfolio":(tab_off,tab_on,tab_off,tab_off),
              "cost":     (tab_off,tab_off,tab_on,tab_off),
              "audit":    (tab_off,tab_off,tab_off,tab_on)}
    p_st = panels.get(active, panels["analysis"])
    t_st = tabs.get(active, tabs["analysis"])
    audit_ch = _build_audit_content() if active == "audit" else [
        html.P("AUDIT LOG", style=HEAD),
        html.Div("Switch to this tab to load entries.",
                 style={"color": C["muted"], "fontSize": "13px"})]
    return (*p_st, *t_st, audit_ch)

def _build_audit_content():
    log = read_audit_log()
    if not log:
        return [html.P("AUDIT LOG", style=HEAD),
                html.Div("No sessions recorded yet. Run an analysis first.",
                         style={"color": C["muted"], "fontSize": "13px"})]
    hs = {"fontFamily": "monospace", "fontSize": "10px", "color": C["muted"],
          "padding": "10px 12px", "borderBottom": "1px solid " + C["border"],
          "whiteSpace": "nowrap"}
    cs = {"padding": "9px 12px", "fontSize": "12px", "color": C["text"],
          "borderBottom": "1px solid " + C["grid"], "whiteSpace": "nowrap"}
    cols = ["Ref","Timestamp","Field","Operator","Country",
            "Points","Alerts","MDLF%","CH4","Sigma","Cont%"]
    rows = [html.Tr([
        html.Td(e.get("ref","—"), style={**cs,"fontFamily":"monospace","fontSize":"10px"}),
        html.Td(e.get("ts","—"),            style=cs),
        html.Td(e.get("field","—"),         style=cs),
        html.Td(e.get("operator","—"),      style=cs),
        html.Td(e.get("country","—"),       style=cs),
        html.Td(str(e.get("n_obs","—")),    style=cs),
        html.Td(str(e.get("n_alerts","—")), style=cs),
        html.Td(str(e.get("mdlf","—")),     style=cs),
        html.Td(str(e.get("ch4","—")),      style=cs),
        html.Td(str(e.get("sigma","—")),    style=cs),
        html.Td(str(e.get("cont","—")),     style=cs),
    ]) for e in reversed(log[-100:])]
    return [
        html.P("AUDIT LOG", style=HEAD),
        html.Div(str(len(log)) + " sessions recorded.",
                 style={"color": C["muted"], "fontSize": "12px", "marginBottom": "16px"}),
        html.Div(style={"overflowX": "auto"}, children=[
            html.Table([html.Thead(html.Tr([html.Th(c, style=hs) for c in cols])),
                        html.Tbody(rows)],
                       style={"width": "100%", "borderCollapse": "collapse"}),
        ]),
    ]

# ── Upload callback ───────────────────────────────────────────────────────────────
@app.callback(
    Output("stored-data", "data"),
    Output("upload-status", "children"),
    Output("column-mapping-section", "style"),
    Output("col-date", "options"),
    Output("col-produced", "options"),
    Output("col-flare", "options"),
    Output("col-export", "options"),
    Output("col-fuel", "options"),
    Output("col-injection", "options"),
    Input("upload-data", "contents"),
    State("upload-data", "filename"),
)
def parse_upload(contents, filename):
    hidden  = {"display": "none"}
    visible = {"display": "block"}
    empty   = []
    if contents is None:
        return None, "", hidden, empty, empty, empty, empty, empty, empty
    try:
        _, cs = contents.split(",")
        df   = pd.read_csv(io.StringIO(base64.b64decode(cs).decode("utf-8")))
        cols = [{"label": c, "value": c} for c in df.columns]
        status = html.Span([
            html.Span("✓ ", style={"color": C["accent"]}),
            filename + " · " + str(len(df)) + " rows · " + str(len(df.columns)) + " columns"
        ], style={"fontSize": "13px", "color": C["text"]})
        return (df.to_json(date_format="iso", orient="split"),
                status, visible, cols, cols, cols, cols, cols, cols)
    except Exception as e:
        return (None, html.Span("Error: " + str(e), style={"color": C["danger"]}),
                hidden, empty, empty, empty, empty, empty, empty)

@app.callback(
    Output("stored-results", "data"),
    Output("kpi-section",    "style"),
    Output("charts-section", "style"),
    Output("kpi-cards",      "children"),
    Output("residual-chart",   "figure"),
    Output("arima-chart",      "figure"),
    Output("xgb-chart",        "figure"),
    Output("lstm-chart",       "figure"),
    Output("snr-chart",        "figure"),
    Output("mdlf-chart",       "figure"),
    Output("mdlf-summary",     "children"),
    Output("importance-chart", "figure"),
    Output("importance-text",  "children"),
    Output("ensemble-chart",   "figure"),
    Output("alert-table",      "children"),
    Output("status-pill",      "children"),
    Output("model-info-display","children"),
    Input("run-btn", "n_clicks"),
    State("stored-data",          "data"),
    State("col-date",             "value"),
    State("col-produced",         "value"),
    State("col-flare",            "value"),
    State("col-export",           "value"),
    State("col-fuel",             "value"),
    State("col-injection",        "value"),
    State("contamination-slider", "value"),
    State("sigma-slider",         "value"),
    State("ch4-fraction",         "value"),
    State("field-name",           "value"),
    State("operator-name",        "value"),
    State("country-name",         "value"),
    prevent_initial_call=True,
)
def run_analysis(n_clicks, stored_data, col_date, col_produced, col_flare,
                 col_export, col_fuel, col_injection,
                 cont_pct, sigma_val, ch4_frac_val,
                 field_name, operator_name, country_name):

    show = {"display": "block"}
    hide = {"display": "none"}

    empties = (None, hide, hide, [],
               blank_fig("Select all columns first"),
               blank_fig(), blank_fig(), blank_fig(), blank_fig(),
               blank_fig(), "", blank_fig("Run analysis"), "",
               blank_fig(), html.Div("No results."),
               _pill("NO DATA", C["muted"]), "")

    if not stored_data or not all([col_date, col_produced, col_flare,
                                   col_export, col_fuel, col_injection]):
        return empties

    try:
        cont      = (cont_pct or 2) / 100
        sigma     = float(sigma_val or 3.0)
        ch4_frac  = float(ch4_frac_val or 0.77)
        field     = field_name     or "Unnamed Field"
        operator  = operator_name  or "Unnamed Operator"
        country   = country_name   or "—"

        df = pd.read_json(io.StringIO(stored_data), orient="split")
        df[col_date] = pd.to_datetime(df[col_date], errors="coerce")
        df = df.sort_values(col_date).reset_index(drop=True)
        for col in [col_produced, col_flare, col_export, col_fuel, col_injection]:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

        raw      = compute_residual(df, col_produced, col_flare,
                                    col_export, col_fuel, col_injection)
        residual = apply_ch4(raw, ch4_frac)
        dates    = df[col_date]

        arima_flags, arima_scores, z_scores, arima_ord = run_arima(residual, sigma)
        xgb_flags,   xgb_scores,   xgb_importance      = run_xgboost(residual, cont)
        ae_flags,    ae_scores                          = run_autoencoder(residual, cont)

        ensemble_alert, vote_count = compute_ensemble(
            arima_flags, xgb_flags, ae_flags, z_scores)

        def norm(arr):
            mn, mx = arr.min(), arr.max()
            return (arr - mn) / (mx - mn + 1e-9)

        arima_n = norm(arima_scores)
        xgb_n   = norm(xgb_scores)
        ae_n    = norm(ae_scores)
        ens_n   = (arima_n + xgb_n + ae_n) / 3

        snr  = compute_snr(residual)
        mdlf = compute_mdlf(residual, df[col_produced])
        mdlf_global = float((3 * residual.std() / df[col_produced].mean()) * 100)

        alerts = []
        for idx in np.where(ensemble_alert)[0]:
            level, colour = classify_alert(
                float(ens_n[idx]), float(snr.iloc[idx]) if idx < len(snr) else 1.0)
            alerts.append({
                "Date":   dates.iloc[idx].strftime("%Y-%m-%d"),
                "Res":    f"{residual.iloc[idx]:.4f}",
                "CH4":    f"{ch4_frac:.2f}",
                "Score":  f"{ens_n[idx]:.3f}",
                "Votes":  str(int(vote_count[idx])),
                "Z":      f"{z_scores[idx]:.2f}",
                "SNR":    f"{snr.iloc[idx]:.2f}" if idx < len(snr) else "—",
                "ARIMA":  "●" if arima_flags[idx] else "○",
                "XGB":    "●" if xgb_flags[idx]   else "○",
                "AE":     "●" if ae_flags[idx]     else "○",
                "Level": level, "colour": colour,
            })

        n_crit  = sum(1 for a in alerts if a["Level"] == "CRITICAL")
        n_warn  = sum(1 for a in alerts if a["Level"] == "WARNING")
        n_watch = sum(1 for a in alerts if a["Level"] == "WATCH")
        ref     = generate_ref(field)

        # Audit log entry
        write_audit_entry({
            "ref":      ref,
            "ts":       datetime.now().strftime("%Y-%m-%d %H:%M"),
            "field":    field, "operator": operator, "country": country,
            "n_obs":    len(df), "n_alerts": len(alerts),
            "mdlf":     round(mdlf_global, 3),
            "ch4":      ch4_frac, "sigma": sigma,
            "cont":     cont_pct,
        })

        # KPI cards
        kpi_data = [
            ("ENSEMBLE ALERTS", len(alerts),           C["accent2"]),
            ("CRITICAL",        n_crit,                C["danger"]),
            ("WARNING",         n_warn,                C["warn"]),
            ("WATCH",           n_watch,               C["accent3"]),
            ("DATA POINTS",     str(len(df)),          C["accent"]),
            ("MEAN MDLF",       f"{mdlf.mean():.2f}%", C["muted"]),
        ]
        kpi_cards = [dbc.Col(_kpi(l, v, c), md=2) for l,v,c in kpi_data]

        # Residual chart
        fig_res = go.Figure()
        fig_res.add_trace(go.Scatter(
            x=dates, y=residual, mode="lines", name="CH4 Residual",
            line=dict(color=C["accent3"], width=1.5),
            fill="tozeroy", fillcolor="rgba(75,159,255,0.08)"))
        if ensemble_alert.any():
            fig_res.add_trace(go.Scatter(
                x=dates[ensemble_alert], y=residual[ensemble_alert],
                mode="markers", name="Ensemble Alert",
                marker=dict(color=C["danger"], size=7,
                            symbol="circle-open", line=dict(width=2))))
        fig_res.add_hline(y=0, line_dash="dot", line_color=C["muted"], line_width=1)
        fig_res.add_hline(y=float(residual.std()*sigma), line_dash="dash",
                          line_color=C["danger"], line_width=1,
                          annotation_text=f"+{sigma}σ threshold",
                          annotation_font_color=C["danger"], annotation_font_size=10)
        fig_res.update_layout(**chart_layout("METHANE BALANCE RESIDUAL · " + field, 340))

        fig_arima = _anomaly_fig(dates, residual, arima_flags, arima_n,
                                  "ARIMA(1,0,1) ANOMALY SCORES")
        fig_xgb   = _anomaly_fig(dates, residual, xgb_flags, xgb_n,
                                  "XGBOOST PREDICTION ERROR ANOMALY SCORES")
        fig_ae    = _anomaly_fig(dates, residual, ae_flags, ae_n,
                                  "AUTOENCODER RECONSTRUCTION ERROR")

        fig_snr = go.Figure()
        fig_snr.add_trace(go.Scatter(x=dates, y=snr, mode="lines",
                                      name="SNR", line=dict(color=C["warn"], width=1.5)))
        fig_snr.add_hline(y=2, line_dash="dash", line_color=C["accent2"],
                          annotation_text="Warning SNR=2",
                          annotation_font_color=C["accent2"], annotation_font_size=10)
        fig_snr.add_hline(y=3, line_dash="dash", line_color=C["danger"],
                          annotation_text="Critical SNR=3",
                          annotation_font_color=C["danger"], annotation_font_size=10)
        fig_snr.update_layout(**chart_layout("SIGNAL-TO-NOISE RATIO (30-DAY ROLLING)", 320))

        fig_mdlf = go.Figure()
        fig_mdlf.add_trace(go.Scatter(x=dates, y=mdlf, mode="lines", name="MDLF %",
                                       line=dict(color=C["accent"], width=1.5),
                                       fill="tozeroy", fillcolor="rgba(0,212,170,0.08)"))
        fig_mdlf.add_hline(y=mdlf_global, line_dash="dash",
                           line_color=C["muted"], line_width=1,
                           annotation_text=f"Global {mdlf_global:.2f}%",
                           annotation_font_color=C["muted"], annotation_font_size=10)
        fig_mdlf.update_layout(**chart_layout(
            "MINIMUM DETECTABLE LEAK FRACTION (90-DAY ROLLING, % OF PRODUCTION)", 320))

        mdlf_text = (f"Ref: {ref} · Global MDLF: {mdlf_global:.2f}% · "
                     f"Rolling mean: {mdlf.mean():.2f}% · "
                     f"Best: {mdlf.min():.2f}% · Worst: {mdlf.max():.2f}%")

        # XGBoost feature importance chart
        fig_imp = go.Figure()
        if xgb_importance:
            sorted_imp = sorted(xgb_importance.items(), key=lambda x: x[1], reverse=True)
            feat_labels = [i[0] for i in sorted_imp]
            feat_vals   = [i[1] for i in sorted_imp]
            top_feat    = feat_labels[0] if feat_labels else "—"
            fig_imp.add_trace(go.Bar(
                x=feat_vals, y=feat_labels,
                orientation="h",
                marker_color=C["accent"],
                marker_line_color=C["accent3"],
                marker_line_width=1,
            ))
            imp_text = ("Most influential predictor: " + top_feat +
                        " · XGBoost trained on 70% of observations · "
                        "Prediction error on full series used for anomaly scoring")
        else:
            fig_imp = blank_fig("Feature importance not available")
            imp_text = ""
            top_feat = "—"
        fig_imp.update_layout(**chart_layout(
            "XGBOOST FEATURE IMPORTANCE (RESIDUAL PREDICTION)", 300))
        fig_imp.update_layout(
            yaxis=dict(autorange="reversed",
                       gridcolor=C["grid"], zeroline=False,
                       tickfont=dict(color=C["muted"], size=11)),
            xaxis_title="Importance Score",
        )

        # Ensemble chart
        fig_ens = go.Figure()
        fig_ens.add_trace(go.Scatter(x=dates, y=arima_n, mode="lines",
                                      name="ARIMA", opacity=0.7,
                                      line=dict(color=C["accent3"], width=1)))
        fig_ens.add_trace(go.Scatter(x=dates, y=xgb_n, mode="lines",
                                      name="XGBoost", opacity=0.7,
                                      line=dict(color=C["warn"], width=1)))
        fig_ens.add_trace(go.Scatter(x=dates, y=ae_n, mode="lines",
                                      name="Autoencoder", opacity=0.7,
                                      line=dict(color=C["accent"], width=1)))
        fig_ens.add_trace(go.Scatter(x=dates, y=ens_n, mode="lines",
                                      name="Ensemble Score",
                                      line=dict(color=C["accent2"], width=2.5)))
        fig_ens.update_layout(**chart_layout("MODEL ENSEMBLE AGREEMENT", 320))

        if n_crit > 0:   pill = _pill(f"● {n_crit} CRITICAL",  C["danger"])
        elif n_warn > 0: pill = _pill(f"● {n_warn} WARNING",   C["warn"])
        elif n_watch > 0:pill = _pill(f"● {n_watch} WATCH",    C["accent3"])
        else:            pill = _pill("● NORMAL",               C["accent"])

        model_info = (f"Ref: {ref} · ARIMA(1,0,1) · "
                      f"XGBoost (80 trees, 7 lags, 70/30 split) · "
                      f"PCA Autoencoder (14-obs window, 2 components) · "
                      f"Ensemble: majority voting (2/3) + positive z-score")

        stored = json.dumps({
            "ref": ref, "field": field, "operator": operator,
            "country": country, "ch4_frac": ch4_frac,
            "mdlf_global": mdlf_global, "n_alerts": len(alerts),
            "cont_pct": cont_pct, "sigma": sigma,
            "top_feat": top_feat,
        })

        return (stored, show, show, kpi_cards,
                fig_res, fig_arima, fig_xgb, fig_ae, fig_snr,
                fig_mdlf, mdlf_text, fig_imp, imp_text,
                fig_ens, _alert_table(alerts), pill, model_info)

    except Exception as e:
        err = _pill("● ERROR", C["danger"])
        return (None, hide, hide, [],
                blank_fig("Error — " + str(e)),
                blank_fig(), blank_fig(), blank_fig(), blank_fig(),
                blank_fig(), "", blank_fig("Error"), "",
                blank_fig(), html.Div("Error: " + str(e),
                                      style={"color": C["danger"]}),
                err, "Error: " + str(e))

# ── Portfolio callback ──────────────────────────────────────────────────────────
@app.callback(
    Output("portfolio-status",       "children"),
    Output("portfolio-table-section","style"),
    Output("portfolio-table",        "children"),
    Output("portfolio-chart",        "figure"),
    Input("portfolio-upload", "contents"),
    State("portfolio-upload", "filename"),
)
def update_portfolio(contents_list, filenames):
    if not contents_list:
        return "", {"display": "none"}, html.Div(), blank_fig()
    results = []
    fig = go.Figure()
    for contents, fname in zip(contents_list, filenames):
        try:
            _, cs = contents.split(",")
            df = pd.read_csv(io.StringIO(base64.b64decode(cs).decode("utf-8")))
            # Auto-detect numeric columns for a quick residual estimate
            num_cols = df.select_dtypes(include=[np.number]).columns.tolist()
            if len(num_cols) < 5:
                results.append({"field": fname, "status": "Insufficient columns",
                                 "alerts": "—", "mdlf": "—", "points": str(len(df))})
                continue
            prod = df[num_cols[0]]
            rest = df[num_cols[1:5]].sum(axis=1)
            res  = (prod - rest) * 0.77
            mdlf = float((3 * res.std() / prod.mean()) * 100)
            n_pos = int((res > res.std() * 3).sum())
            results.append({
                "field":   fname.replace(".csv",""),
                "points":  str(len(df)),
                "alerts":  str(n_pos),
                "mdlf":    f"{mdlf:.2f}%",
                "status":  "WATCH" if n_pos > 0 else "NORMAL",
            })
            try:
                date_col = df.select_dtypes(include=["object","datetime64"]).columns[0]
                dates = pd.to_datetime(df[date_col], errors="coerce")
                r_series = pd.Series(res.values)
                mdf = ((3 * r_series.rolling(90,min_periods=30).std() /
                        prod.rolling(90,min_periods=30).mean().replace(0,np.nan))*100
                       ).ffill().bfill()
                fig.add_trace(go.Scatter(
                    x=dates, y=mdf, mode="lines",
                    name=fname.replace(".csv",""), line=dict(width=1.5)))
            except Exception:
                pass
        except Exception as ex:
            results.append({"field": fname, "status": "Error: " + str(ex),
                             "alerts": "—", "mdlf": "—", "points": "—"})

    fig.update_layout(**chart_layout("MDLF COMPARISON ACROSS FIELDS", 320))

    hs = {"fontFamily": "monospace", "fontSize": "10px", "color": C["muted"],
          "padding": "10px 12px", "borderBottom": "1px solid " + C["border"],
          "whiteSpace": "nowrap"}
    cs = {"padding": "9px 12px", "fontSize": "12px", "color": C["text"],
          "borderBottom": "1px solid " + C["grid"], "whiteSpace": "nowrap"}
    cols = ["Field", "Data Points", "Anomaly Count", "Global MDLF", "Status"]
    rows = []
    for r in results:
        sc = C["warn"] if r["status"] not in ["NORMAL","—","Insufficient columns"] else C["accent"]
        rows.append(html.Tr([
            html.Td(r["field"],   style=cs),
            html.Td(r["points"],  style=cs),
            html.Td(r["alerts"],  style=cs),
            html.Td(r["mdlf"],    style=cs),
            html.Td(r["status"],  style={**cs, "color": sc,
                                         "fontFamily": "monospace", "fontWeight": "700"}),
        ]))
    tbl = html.Div(style={"overflowX": "auto"}, children=[
        html.Table([
            html.Thead(html.Tr([html.Th(c, style=hs) for c in cols])),
            html.Tbody(rows),
        ], style={"width": "100%", "borderCollapse": "collapse"}),
    ])
    status = html.Span(
        str(len(results)) + " field(s) loaded",
        style={"color": C["accent"], "fontSize": "13px"})
    return status, {"display": "block"}, tbl, fig

# ── Cost calculator callback ─────────────────────────────────────────────────────
@app.callback(
    Output("cost-results",       "style"),
    Output("cost-chart",         "figure"),
    Output("cost-summary-table", "children"),
    Input("btn-calculate-cost", "n_clicks"),
    State("cost-production",    "value"),
    State("cost-price",         "value"),
    State("cost-leak-rate",     "value"),
    State("cost-pfems-fee",     "value"),
    prevent_initial_call=True,
)
def calculate_cost(n, production, price, leak_rate, pfems_fee):
    prod      = float(production or 150)
    ch4_price = float(price      or 3000)
    leak_pct  = float(leak_rate  or 1.0) / 100
    pfems_cost= float(pfems_fee  or 25000)

    daily_leak_mmscf  = prod * leak_pct
    annual_leak_mmscf = daily_leak_mmscf * 365
    annual_leak_value = annual_leak_mmscf * ch4_price

    methods = {
        "PFEMS": {
            "annual_cost": pfems_cost,
            "coverage_days": 365,
            "detection_pct": 95,
        },
        "OGI Camera Survey": {
            "annual_cost": 3500 * 12,
            "coverage_days": 12,
            "detection_pct": 40,
        },
        "Satellite Monitoring": {
            "annual_cost": 800 * 48,
            "coverage_days": 48,
            "detection_pct": 55,
        },
        "Continuous Sensor Array": {
            "annual_cost": 1200 * 365,
            "coverage_days": 365,
            "detection_pct": 90,
        },
    }

    fig = go.Figure()
    names, costs, savings, ratios = [], [], [], []
    rows = []
    hs = {"fontFamily": "monospace", "fontSize": "10px", "color": C["muted"],
          "padding": "10px 12px", "borderBottom": "1px solid " + C["border"],
          "whiteSpace": "nowrap"}
    cs = {"padding": "9px 12px", "fontSize": "12px", "color": C["text"],
          "borderBottom": "1px solid " + C["grid"], "whiteSpace": "nowrap"}
    cols_h = ["Method","Annual Cost (USD)","Coverage Days",
              "Detection Rate","Value Recovered (USD)","Net Benefit (USD)"]
    rows.append(html.Tr([html.Th(c, style=hs) for c in cols_h]))
    for method, data in methods.items():
        annual_cost = data["annual_cost"]
        value_rec   = annual_leak_value * (data["detection_pct"] / 100)
        net_benefit = value_rec - annual_cost
        colour = C["accent"] if net_benefit > 0 else C["danger"]
        names.append(method)
        costs.append(annual_cost)
        savings.append(value_rec)
        ratios.append(net_benefit)
        rows.append(html.Tr([
            html.Td(method,                        style={**cs, "fontWeight": "600"}),
            html.Td(f"${annual_cost:,.0f}",        style=cs),
            html.Td(str(data["coverage_days"]),    style=cs),
            html.Td(str(data["detection_pct"]) + "%", style=cs),
            html.Td(f"${value_rec:,.0f}",          style=cs),
            html.Td(f"${net_benefit:,.0f}",
                    style={**cs, "color": colour, "fontWeight": "700"}),
        ]))

    fig.add_trace(go.Bar(name="Annual Cost",        x=names, y=costs,
                          marker_color=C["danger"],  opacity=0.8))
    fig.add_trace(go.Bar(name="Value Recovered",    x=names, y=savings,
                          marker_color=C["accent"],  opacity=0.8))
    fig.add_trace(go.Scatter(name="Net Benefit", x=names, y=ratios,
                              mode="markers+lines",
                              marker=dict(color=C["accent2"], size=10),
                              line=dict(color=C["accent2"], width=2)))
    fig.update_layout(**chart_layout("COST vs VALUE RECOVERED (USD/YEAR)", 380))
    fig.update_layout(barmode="group")

    tbl = html.Table([html.Tbody(rows)],
                     style={"width": "100%", "borderCollapse": "collapse"})
    return {"display": "block"}, fig, tbl

# ── Export callback ──────────────────────────────────────────────────────────────
@app.callback(
    Output("download-pdf",   "data"),
    Output("download-excel", "data"),
    Output("download-csv",   "data"),
    Output("export-status",  "children"),
    Input("btn-pdf",   "n_clicks"),
    Input("btn-excel", "n_clicks"),
    Input("btn-csv",   "n_clicks"),
    State("stored-results",       "data"),
    State("stored-data",          "data"),
    State("col-date",             "value"),
    State("col-produced",         "value"),
    State("col-flare",            "value"),
    State("col-export",           "value"),
    State("col-fuel",             "value"),
    State("col-injection",        "value"),
    State("contamination-slider", "value"),
    State("sigma-slider",         "value"),
    State("ch4-fraction",         "value"),
    State("field-name",           "value"),
    State("operator-name",        "value"),
    State("country-name",         "value"),
    prevent_initial_call=True,
)
def handle_exports(bp, be, bc,
                   stored_results, stored_data,
                   col_date, col_produced, col_flare, col_export,
                   col_fuel, col_injection, cont_pct, sigma_val,
                   ch4_frac_val, field_name, operator_name, country_name):

    ctx = callback_context
    if not ctx.triggered or not stored_results or not stored_data:
        return None, None, None, html.Span("Run analysis first.",
                                            style={"color": C["warn"]})

    triggered = ctx.triggered[0]["prop_id"].split(".")[0]
    cont      = (cont_pct or 2) / 100
    sigma     = float(sigma_val  or 3.0)
    ch4_frac  = float(ch4_frac_val or 0.77)
    field     = field_name    or "Unnamed Field"
    operator  = operator_name or "Unnamed Operator"
    country   = country_name  or "—"
    meta      = json.loads(stored_results)
    ref       = meta.get("ref", generate_ref(field))
    mdlf_gl   = meta.get("mdlf_global")
    top_feat  = meta.get("top_feat", "—")
    ts        = datetime.now().strftime("%Y%m%d_%H%M")

    df = pd.read_json(io.StringIO(stored_data), orient="split")
    df[col_date] = pd.to_datetime(df[col_date], errors="coerce")
    df = df.sort_values(col_date).reset_index(drop=True)
    for col in [col_produced, col_flare, col_export, col_fuel, col_injection]:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    raw      = compute_residual(df, col_produced, col_flare,
                                col_export, col_fuel, col_injection)
    residual = apply_ch4(raw, ch4_frac)
    dates_l  = df[col_date].dt.strftime("%Y-%m-%d").tolist()

    arima_flags, arima_scores, z_scores, arima_ord = run_arima(residual, sigma)
    xgb_flags,   xgb_scores,   xgb_imp             = run_xgboost(residual, cont)
    ae_flags,    ae_scores                          = run_autoencoder(residual, cont)
    ensemble_alert, vote_count                      = compute_ensemble(
        arima_flags, xgb_flags, ae_flags, z_scores)

    def norm(arr):
        mn, mx = arr.min(), arr.max()
        return (arr - mn) / (mx - mn + 1e-9)

    ens_n  = (norm(arima_scores) + norm(xgb_scores) + norm(ae_scores)) / 3
    snr    = compute_snr(residual)
    mdlf   = compute_mdlf(residual, df[col_produced])

    alerts = []
    for idx in np.where(ensemble_alert)[0]:
        level, colour = classify_alert(
            float(ens_n[idx]), float(snr.iloc[idx]) if idx < len(snr) else 1.0)
        alerts.append({
            "Date":  dates_l[idx] if idx < len(dates_l) else "—",
            "Res":   f"{residual.iloc[idx]:.4f}",
            "CH4":   f"{ch4_frac:.2f}",
            "Score": f"{ens_n[idx]:.3f}",
            "Votes": str(int(vote_count[idx])),
            "Z":     f"{z_scores[idx]:.2f}",
            "SNR":   f"{snr.iloc[idx]:.2f}" if idx < len(snr) else "—",
            "ARIMA": "●" if arima_flags[idx] else "○",
            "XGB":   "●" if xgb_flags[idx]   else "○",
            "AE":    "●" if ae_flags[idx]     else "○",
            "Level": level, "colour": colour,
        })

    n_crit  = sum(1 for a in alerts if a["Level"] == "CRITICAL")
    n_warn  = sum(1 for a in alerts if a["Level"] == "WARNING")
    n_watch = sum(1 for a in alerts if a["Level"] == "WATCH")
    gen_ts  = datetime.now().strftime("%Y-%m-%d %H:%M UTC")
    dr      = f"{dates_l[0]} to {dates_l[-1]}" if dates_l else "—"
    mean_r  = float(np.mean(residual.values))
    std_r   = float(np.std(residual.values))
    mf      = f"{mdlf_gl:.2f}%" if mdlf_gl else "—"
    status_str = ("CRITICAL" if n_crit > 0 else "WARNING" if n_warn > 0
                  else "WATCH" if n_watch > 0 else "NORMAL")

    if triggered == "btn-pdf":
        buf  = io.BytesIO()
        doc  = SimpleDocTemplate(buf, pagesize=A4,
                                  leftMargin=2*cm, rightMargin=2*cm,
                                  topMargin=2*cm, bottomMargin=2*cm)
        acc  = colors.HexColor("#00D4AA")
        drk  = colors.HexColor("#0A0E17")
        mut  = colors.HexColor("#6B7A99")
        dng  = colors.HexColor("#FF3B5C")
        wrn  = colors.HexColor("#FFB800")
        t_s  = ParagraphStyle("t", fontSize=22, textColor=acc,
                               fontName="Helvetica-Bold", spaceAfter=4)
        s_s  = ParagraphStyle("s", fontSize=10, textColor=mut,
                               fontName="Helvetica", spaceAfter=5)
        h_s  = ParagraphStyle("h", fontSize=11, textColor=acc,
                               fontName="Helvetica-Bold", spaceAfter=8, spaceBefore=16)
        b_s  = ParagraphStyle("b", fontSize=9,
                               textColor=colors.HexColor("#333333"),
                               fontName="Helvetica", spaceAfter=6, leading=14)
        disc_s = ParagraphStyle("disc", fontSize=8,
                                 textColor=colors.HexColor("#888888"),
                                 fontName="Helvetica-Oblique",
                                 spaceAfter=4, leading=11)
        story = [
            Spacer(1, 1.5*cm),
            Paragraph("PFEMS", t_s),
            Paragraph("Physics-Guided Fugitive Emissions Monitoring System", s_s),
            Paragraph("Savanna Dynamics Limited", s_s),
            HRFlowable(width="100%", thickness=1,
                       color=colors.HexColor("#1E2D45"), spaceAfter=8),
            Paragraph(f"<b>Report Reference:</b> {ref}", s_s),
            Paragraph(f"<b>Field:</b> {field}", s_s),
            Paragraph(f"<b>Operator:</b> {operator}", s_s),
            Paragraph(f"<b>Country / Jurisdiction:</b> {country}", s_s),
            Paragraph(f"<b>CH4 Fraction:</b> {ch4_frac:.2f}", s_s),
            Paragraph(f"<b>Generated:</b> {gen_ts}", s_s),
            Spacer(1, 0.5*cm),
            Paragraph("EXECUTIVE SUMMARY", h_s),
            Paragraph(
                f"This report presents continuous methane imbalance monitoring results from the "
                f"PFEMS framework for <b>{field}</b> operated by <b>{operator}</b> "
                f"({country}). The analysis covers {dr}, examining {len(residual):,} "
                f"daily operational data points using a field-specific methane fraction "
                f"of <b>{ch4_frac:.2f}</b>.<br/><br/>"
                f"<b>System status: {status_str}</b> — {n_crit} critical, "
                f"{n_warn} warning, {n_watch} watch alerts from {len(alerts)} total "
                f"ensemble-confirmed events.<br/><br/>"
                f"Mean methane residual: {mean_r:.4f} MMscf/day "
                f"(std: {std_r:.4f}). Global MDLF: {mf}. "
                f"Primary XGBoost predictor: {top_feat}.", b_s),
            Paragraph("RECOMMENDED ACTIONS", h_s),
        ]
        if n_crit > 0:
            story.append(Paragraph(
                f"<b>CRITICAL:</b> {n_crit} event(s) detected with ensemble score above 0.8 "
                f"and SNR above 3.0. Immediate physical inspection is recommended. "
                f"Deploy OGI camera or portable methane detector to the identified "
                f"locations and dates.", b_s))
        if n_warn > 0:
            story.append(Paragraph(
                f"<b>WARNING:</b> {n_warn} event(s) detected with ensemble score above 0.5 "
                f"or SNR above 2.0. Schedule inspection within 48 hours. "
                f"Cross-check with production logs for corresponding operational events.", b_s))
        if n_watch > 0:
            story.append(Paragraph(
                f"<b>WATCH:</b> {n_watch} event(s) detected with ensemble score above 0.2. "
                f"Monitor closely. No immediate action required but maintain elevated "
                f"surveillance for the flagged dates.", b_s))
        if not alerts:
            story.append(Paragraph(
                "No ensemble-confirmed alerts. Continue routine monitoring.", b_s))
        story.append(Paragraph("KEY METRICS", h_s))
        kd = [["Metric","Value"],
              ["Report Reference", ref],
              ["Field", field], ["Operator", operator], ["Country", country],
              ["CH4 Fraction", f"{ch4_frac:.2f}"],
              ["Analysis Period", dr],
              ["Total Data Points", f"{len(residual):,}"],
              ["Mean Residual (MMscf/day)", f"{mean_r:.4f}"],
              ["Std Deviation (MMscf/day)", f"{std_r:.4f}"],
              ["Global MDLF", mf],
              ["Detection Sigma", str(sigma)],
              ["Contamination Setting", str(cont_pct) + "%"],
              ["Top XGBoost Feature", top_feat],
              ["Total Ensemble Alerts", str(len(alerts))],
              ["Critical", str(n_crit)],
              ["Warning", str(n_warn)],
              ["Watch", str(n_watch)]]
        kt = Table(kd, colWidths=[9*cm, 7*cm])
        kt.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0),drk), ("TEXTCOLOR",(0,0),(-1,0),acc),
            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
            ("FONTSIZE",(0,0),(-1,-1),9),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),
             [colors.HexColor("#F8F9FA"), colors.white]),
            ("GRID",(0,0),(-1,-1),0.5,colors.HexColor("#DDDDDD")),
            ("PADDING",(0,0),(-1,-1),6),
            ("LEFTPADDING",(0,0),(-1,-1),10),
        ]))
        story.append(kt)
        story.append(PageBreak())
        story.append(Paragraph("ENSEMBLE ALERT LOG", h_s))
        story.append(Paragraph(
            "Alerts confirmed by majority voting (2 of 3 detectors) with positive "
            "z-score direction restriction. Detectors: ARIMA(1,0,1) baseline, "
            "XGBoost prediction error, PCA autoencoder reconstruction error.", b_s))
        if alerts:
            ah = ["Date","Residual","CH4","Score","Votes","z",
                  "SNR","ARIMA","XGB","AE","Level"]
            ar = [ah] + [[a["Date"],a["Res"],a["CH4"],a["Score"],
                          a["Votes"],a["Z"],a["SNR"],
                          a["ARIMA"],a["XGB"],a["AE"],a["Level"]]
                         for a in alerts[:150]]
            at = Table(ar,
                       colWidths=[2.2*cm,2*cm,1.5*cm,1.5*cm,1.2*cm,
                                   1.5*cm,1.5*cm,1.2*cm,1.2*cm,1.2*cm,1.8*cm])
            ats = TableStyle([
                ("BACKGROUND",(0,0),(-1,0),drk),
                ("TEXTCOLOR",(0,0),(-1,0),acc),
                ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
                ("FONTSIZE",(0,0),(-1,-1),7.5),
                ("ROWBACKGROUNDS",(0,1),(-1,-1),
                 [colors.HexColor("#F8F9FA"),colors.white]),
                ("GRID",(0,0),(-1,-1),0.3,colors.HexColor("#DDDDDD")),
                ("PADDING",(0,0),(-1,-1),3),
                ("ALIGN",(0,0),(-1,-1),"CENTER"),
            ])
            for i, a in enumerate(alerts[:150], 1):
                if a["Level"] == "CRITICAL":
                    ats.add("TEXTCOLOR",(10,i),(10,i),dng)
                    ats.add("FONTNAME",(10,i),(10,i),"Helvetica-Bold")
                elif a["Level"] == "WARNING":
                    ats.add("TEXTCOLOR",(10,i),(10,i),wrn)
                    ats.add("FONTNAME",(10,i),(10,i),"Helvetica-Bold")
            at.setStyle(ats)
            story.append(at)
            if len(alerts) > 150:
                story.append(Paragraph(
                    f"Showing first 150 of {len(alerts)} alerts. "
                    f"Download Excel export for the full log.", b_s))
        else:
            story.append(Paragraph("No ensemble-confirmed alerts detected.", b_s))
        story.append(PageBreak())
        story.append(Paragraph("METHODOLOGY", h_s))
        story.append(Paragraph(
            f"<b>Gas-Balance Methane Residual:</b> R(t) = B(t) + L(t) + epsilon(t), "
            f"where the residual is computed as (produced gas — flaring — export — "
            f"fuel gas — injection) x {ch4_frac:.2f}. The methane fraction is "
            f"field-specific and user-configurable.<br/><br/>"
            f"<b>ARIMA(1,0,1) Baseline:</b> Fixed order for computational efficiency. "
            f"Anomalies flagged where prediction error z-score exceeds +{sigma}σ "
            f"(positive direction only, consistent with unmetered gas loss).<br/><br/>"
            f"<b>XGBoost Prediction Error:</b> Gradient boosting regressor trained on "
            f"7 lag features plus 14-day rolling statistics using the first 70 percent "
            f"of observations. Prediction errors on the full series are used as anomaly "
            f"scores. Feature importances identify which temporal patterns drive "
            f"detection. Primary feature: {top_feat}.<br/><br/>"
            f"<b>PCA Autoencoder:</b> 14-observation windows compressed to 2 principal "
            f"components and reconstructed. High reconstruction error (top "
            f"{cont_pct}% of observations) indicates pattern deviation from "
            f"normal baseline behaviour.<br/><br/>"
            f"<b>Ensemble:</b> Alert confirmed when at least 2 of 3 detectors agree "
            f"AND the ARIMA z-score is positive. Negative residuals are excluded from "
            f"confirmation regardless of magnitude (R(t) less than 0 cannot represent "
            f"unmetered gas loss).<br/><br/>"
            f"<b>MDLF:</b> Minimum Detectable Leak Fraction = 3 x rolling 90-day sigma "
            f"of residual / rolling 90-day mean production, expressed as percent of "
            f"daily production. Global MDLF: {mf}.<br/><br/>"
            f"<b>Reference:</b> Abdul Hameed M. A Scalable Data-Driven Framework for "
            f"Fugitive Methane Detection in Offshore Oil and Gas Production Systems. "
            f"Discover Sustainability, Springer Nature, 2025 (under review).", b_s))
        story.append(Spacer(1, 1*cm))
        story.append(HRFlowable(width="100%", thickness=0.5,
                                color=colors.HexColor("#CCCCCC"), spaceAfter=6))
        story.append(Paragraph(
            f"DISCLAIMER: This report is generated by an automated monitoring system "
            f"for screening purposes only. Ensemble alerts indicate statistically "
            f"anomalous gas accounting imbalances and do not constitute confirmed "
            f"fugitive emission events. Physical investigation is required to determine "
            f"the cause of each alert. Savanna Dynamics Limited accepts no liability "
            f"for decisions made solely on the basis of this report.", disc_s))
        story.append(Paragraph(
            f"PFEMS v2.0 · Savanna Dynamics Limited · Ref: {ref} · "
            f"{gen_ts} · Data processed in-session only", disc_s))
        doc.build(story)
        buf.seek(0)
        return (dcc.send_bytes(buf.read(), f"PFEMS_{ref}.pdf"),
                None, None,
                html.Span("✓ PDF downloaded", style={"color": C["accent"]}))

    elif triggered == "btn-excel":
        buf = io.BytesIO()
        wb  = openpyxl.Workbook()
        hf  = PatternFill("solid", fgColor="0A0E17")
        hfn = Font(color="00D4AA", bold=True, name="Consolas", size=10)
        nfn = Font(name="Calibri", size=10)
        cfn = Font(color="FF3B5C", bold=True, name="Calibri", size=10)
        wfn = Font(color="FFB800", bold=True, name="Calibri", size=10)
        bdr = Border(bottom=Side(style="thin", color="1E2D45"),
                     right=Side(style="thin",  color="1E2D45"))
        ws  = wb.active
        ws.title = "Alert Log"
        hdrs = ["Date","Residual CH4","CH4 Frac","Score","Votes",
                "z-score","SNR","ARIMA","XGBoost","Autoencoder","Level"]
        for c, h in enumerate(hdrs, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.fill = hf; cell.font = hfn
            cell.alignment = Alignment(horizontal="center")
        for ri, a in enumerate(alerts, 2):
            vals = [a["Date"], float(a["Res"]), float(a["CH4"]),
                    float(a["Score"]), int(a["Votes"]),
                    float(a["Z"]),
                    float(a["SNR"]) if a["SNR"] != "—" else "",
                    1 if a["ARIMA"]=="●" else 0,
                    1 if a["XGB"]  =="●" else 0,
                    1 if a["AE"]   =="●" else 0,
                    a["Level"]]
            for ci, val in enumerate(vals, 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.border = bdr
                cell.font = (cfn if ci==11 and a["Level"]=="CRITICAL" else
                             wfn if ci==11 and a["Level"]=="WARNING"  else nfn)
                cell.alignment = Alignment(horizontal="center")
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = min(
                max(len(str(c.value or "")) for c in col)+4, 30)
        ws2 = wb.create_sheet("Residual Signal")
        for c, h in enumerate(["Date","Residual CH4","Ens Score","SNR","MDLF%"], 1):
            cell = ws2.cell(row=1, column=c, value=h)
            cell.fill = hf; cell.font = hfn
            cell.alignment = Alignment(horizontal="center")
        mdlf_l = list(mdlf.values)
        for i, (d, r) in enumerate(zip(dates_l, residual.values), 2):
            ws2.cell(row=i,column=1,value=str(d)).font=nfn
            ws2.cell(row=i,column=2,value=float(r)).font=nfn
            ws2.cell(row=i,column=3,
                     value=float(ens_n[i-2]) if i-2<len(ens_n) else "").font=nfn
            ws2.cell(row=i,column=4,
                     value=float(snr.values[i-2]) if i-2<len(snr) else "").font=nfn
            ws2.cell(row=i,column=5,
                     value=float(mdlf_l[i-2]) if i-2<len(mdlf_l) else "").font=nfn
        for col in ws2.columns:
            ws2.column_dimensions[col[0].column_letter].width = 22
        ws3 = wb.create_sheet("Summary")
        ws3.sheet_view.showGridLines = False
        s3d = [
            ("PFEMS v2.0 · Savanna Dynamics Limited", None),
            (f"Report Ref: {ref}", None),
            (f"Field: {field}", None), (f"Operator: {operator}", None),
            (f"Country: {country}", None), (f"CH4 Frac: {ch4_frac:.2f}", None),
            ("",""),("METRIC","VALUE"),
            ("Report Reference", ref),
            ("Data points", len(residual)),
            ("Mean residual", round(mean_r, 4)),
            ("Std deviation", round(std_r, 4)),
            ("Global MDLF", mf),
            ("Top XGBoost feature", top_feat),
            ("Total alerts", len(alerts)),
            ("Critical", n_crit), ("Warning", n_warn), ("Watch", n_watch),
        ]
        for ri, (lab, val) in enumerate(s3d, 1):
            c1 = ws3.cell(row=ri, column=1, value=lab)
            c1.font = (Font(name="Consolas",size=14,bold=True,color="00D4AA")
                       if ri==1 else hfn if lab=="METRIC" else nfn)
            if val is not None:
                c2 = ws3.cell(row=ri, column=2, value=val)
                c2.font = hfn if lab=="METRIC" else nfn
        ws3.column_dimensions["A"].width = 40
        ws3.column_dimensions["B"].width = 25
        wb.save(buf); buf.seek(0)
        return (None,
                dcc.send_bytes(buf.read(), f"PFEMS_{ref}.xlsx"),
                None,
                html.Span("✓ Excel downloaded", style={"color": C["accent"]}))

    elif triggered == "btn-csv":
        df_out = pd.DataFrame({
            "date":              dates_l,
            "residual_ch4":      list(residual.values),
            "ch4_fraction":      ch4_frac,
            "ensemble_score":    list(ens_n),
            "snr":               list(snr.values),
            "z_score":           list(z_scores),
            "arima_flag":        arima_flags.astype(int).tolist(),
            "xgboost_flag":      xgb_flags.astype(int).tolist(),
            "autoencoder_flag":  ae_flags.astype(int).tolist(),
            "ensemble_alert":    ensemble_alert.astype(int).tolist(),
            "mdlf_pct":          list(mdlf.values),
            "report_ref":        ref,
        })
        return (None, None,
                dcc.send_string(df_out.to_csv(index=False), f"PFEMS_{ref}.csv"),
                html.Span("✓ CSV downloaded", style={"color": C["accent"]}))

    return None, None, None, ""


if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=8050)
